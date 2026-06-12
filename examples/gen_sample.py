# -*- coding: utf-8 -*-
"""
示例脚本：生成「高级产品经理」面试打分表
用途：作为 lark-hr-interview Skill 的可见产物样例
运行：python examples/gen_sample.py
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

# ── 样式辅助函数 ──
def hfill(h): return PatternFill("solid", fgColor=h)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def wa(h="left", v="top"): return Alignment(wrap_text=True, horizontal=h, vertical=v)
def hdr(ws, row, cols, texts, bg="1F4E79", fg="FFFFFF"):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = hfill(bg); c.font = Font(bold=True, size=10, color=fg)
        c.alignment = wa("center","center"); c.border = thin_border()
def cell(ws, row, col, val, bg=None, bold=False, h="left"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = hfill(bg)
    c.font = Font(bold=bold, size=10)
    c.alignment = wa(h, "top"); c.border = thin_border()
    return c

# ── 数据 ──
JOB_TITLE = "高级产品经理"
JOB_DEPT  = "增长事业部"
JOB_DATE  = str(date.today())

ABILITIES = [
    ("C1", "需求洞察力", "能从用户反馈、数据和市场信号中精准识别真实需求，区分需求与伪需求"),
    ("C2", "产品规划能力", "能制定清晰的产品路线图，平衡短期交付与长期战略，并能向团队清晰传达"),
    ("C3", "跨职能协作", "能高效推动研发、设计、运营等多团队对齐，处理优先级冲突，确保按时交付"),
    ("C4", "数据驱动决策", "能设计合理的指标体系，用数据验证假设，基于数据做产品迭代判断"),
]
TRAITS = [
    ("T1", "用户同理心", "真正关注用户体验，能站在用户视角发现产品问题，而非只关注业务指标"),
    ("T2", "抗压与推进力", "在资源不足、需求频繁变动的环境下，仍能保持节奏推动项目落地"),
]

QUESTIONS = [
    dict(id="Q01", dim="需求洞察力", dt="核心能力", tp="行为题",
         q="请分享一次你识别出一个被团队忽视的用户真实需求的经历。你是如何发现的，最终如何推动它落地？",
         s1="无法描述具体经历，或只能复述表面现象",
         s2="有经历但主要依赖他人指出，自身发现能力弱",
         s3="能描述系统性的信息收集方式（用研/数据/竞品），有明确结论",
         s4="有完整的发现→验证→推动链路，有具体数据佐证",
         s5="发现的需求带来显著业务结果（DAU/转化/NPS等可量化提升），并形成可复用的方法论",
         sS="当时的业务背景是什么？团队规模和所处阶段？",
         sT="你在这件事中具体负责什么？目标是什么？",
         sA="你用了哪些方法来发现和验证这个需求？",
         sR="最终需求落地了吗？带来了什么结果？有数据吗？",
         sX="如果当时资源更有限，你会优先做什么？"),
    dict(id="Q02", dim="需求洞察力", dt="核心能力", tp="情景题",
         q="假设用户反馈显示某核心功能满意度持续下滑，但数据显示使用率没有明显变化。你会如何判断真实问题并决定下一步行动？",
         s1="只凭感觉或直接按反馈字面意思处理",
         s2="知道要分析数据，但思路停留在表面",
         s3="能拆解满意度与使用率背离的可能原因（沉默用户、替代路径等），提出验证方案",
         s4="有完整的假设树 + 验证优先级 + 快速实验方案，考虑资源约束",
         s5="能主动识别潜在风险（如沉默流失），提出超出问题本身的系统性改进方案",
         sS="这个功能在业务中处于什么位置？用户规模多大？",
         sT="你作为 PM 在这个场景下的决策权边界是什么？",
         sA="你会设计什么样的快速验证实验？",
         sR="你判断最可能的原因是什么？依据是？",
         sX="如果验证周期需要 4 周，但业务方只给你 1 周，你怎么办？"),
    dict(id="Q03", dim="产品规划能力", dt="核心能力", tp="行为题",
         q="请描述你主导制定过的一次产品路线图。你是如何平衡不同团队的优先级诉求，并最终让大家对路线图达成共识的？",
         s1="无法描述具体过程，或路线图由上级指定",
         s2="有参与但主要执行，缺乏自主判断",
         s3="能描述路线图制定的基本流程（收集诉求→评估优先级→对齐），有清晰输出",
         s4="有明确的优先级框架（如 ICE/RICE），能处理跨团队冲突，路线图被实际执行",
         s5="路线图带来显著业务结果，且方法论被团队复用或推广",
         sS="这是什么产品/业务线的路线图？当时有哪些主要干系人？",
         sT="你的决策权边界在哪里？有哪些约束条件？",
         sA="你用什么框架来评估和排序不同需求？",
         sR="路线图执行情况如何？哪些做到了，哪些没做到？",
         sX="回顾来看，有没有一个决策你现在会做不同的选择？"),
    dict(id="Q04", dim="跨职能协作", dt="核心能力", tp="行为题",
         q="请分享一次你在项目中遇到跨团队严重分歧（如研发评估工期远超预期，或设计方案与业务目标冲突）时的处理经历。",
         s1="主要依赖上级介入解决，自身回避冲突",
         s2="能推动对话，但结果靠妥协而非真正对齐",
         s3="能识别分歧根源，组织有效沟通，找到可接受方案",
         s4="有清晰的冲突处理框架，能保持各方关系的同时推动项目，有具体结果",
         s5="处理方式被团队认可并形成协作规范，项目成功交付且超出预期",
         sS="分歧发生在什么背景下？涉及哪些团队？",
         sT="你在这个项目中的角色和责任是什么？",
         sA="你采取了哪些具体步骤来解决分歧？",
         sR="最终结果怎样？项目有没有按时交付？",
         sX="这次经历让你对跨团队协作有什么新的认知？"),
    dict(id="Q05", dim="数据驱动决策", dt="核心能力", tp="情景题",
         q="你上线了一个新功能，上线后核心指标 A 提升了 5%，但另一个重要指标 B 下降了 3%。你会如何分析和决策？",
         s1="只关注正向指标，忽略负向影响",
         s2="注意到 B 下降，但不知道如何系统分析",
         s3="能拆解 A/B 的关联关系，评估净影响，提出进一步验证方向",
         s4="有完整的影响评估框架（用户分层/漏斗分析/长短期影响），给出清晰的保留/回滚/迭代建议",
         s5="能识别指标背后的业务逻辑冲突，提出同时优化 A 和 B 的系统方案，并设计后续监控机制",
         sS="A 和 B 分别是什么性质的指标？对业务的相对重要性如何？",
         sT="你有多大的决策权？需要汇报给谁？",
         sA="你会优先做哪些分析？用什么工具？",
         sR="你倾向于什么决策，依据是什么？",
         sX="如果 A 是 CEO 最关注的指标，但你判断 B 的下降更危险，你会怎么做？"),
    dict(id="Q06", dim="用户同理心", dt="关键特质", tp="行为题",
         q="请分享一次你因为真正站在用户视角，推翻了原本被团队认可的产品方案的经历。",
         s1="无相关经历，或只是表面认同用户视角",
         s2="有经历但主要是配合用研结论，非主动发现",
         s3="能描述具体的用户洞察过程，以及如何说服团队调整方案",
         s4="有清晰的用户问题识别 → 方案质疑 → 替代方案提出 → 落地的完整链路",
         s5="方案调整带来显著的用户体验或业务结果提升，且推动团队建立了更好的用户视角决策文化",
         sS="原方案是什么？为什么最初被认可？",
         sT="你在这件事中的角色是什么？",
         sA="你是如何发现用户真实问题的？用了什么方法？",
         sR="最终方案调整了吗？结果如何？",
         sX="如果团队坚持原方案，你会怎么办？"),
    dict(id="Q07", dim="抗压与推进力", dt="关键特质", tp="情景题",
         q="假设你正在推进的核心项目在上线前 2 周，关键研发资源被临时抽调支援另一个紧急项目，导致交付面临严重风险。你会如何应对？",
         s1="等待上级解决，或直接接受延期",
         s2="能上报问题，但缺乏主动的解决方案",
         s3="能快速评估影响，提出范围裁剪或资源申请的备选方案",
         s4="有清晰的风险评估 + 多套应对预案 + 主动的干系人沟通，保持项目可控",
         s5="在极限约束下仍能保证核心功能交付，并将这次经历转化为团队的应急预案",
         sS="这个项目对业务的重要性如何？当时的时间和资源约束是什么？",
         sT="你的决策权边界在哪里？",
         sA="你会优先采取哪些行动？",
         sR="你认为最可能的结果是什么？",
         sX="如果无论如何都必须延期，你会如何向业务方解释并重建信任？"),
]

# ── 创建工作簿 ──
wb = openpyxl.Workbook()

# Sheet1: 封面
ws1 = wb.active; ws1.title = "封面"
ws1.column_dimensions["A"].width = 16
ws1.column_dimensions["B"].width = 62
ws1.merge_cells("A1:B1")
c = ws1["A1"]; c.value = "HR 面试题设计工作台"
c.fill = hfill("1F4E79"); c.font = Font(bold=True, size=16, color="FFFFFF")
c.alignment = Alignment(horizontal="center", vertical="center"); ws1.row_dimensions[1].height = 42
ws1.merge_cells("A2:B2")
c = ws1["A2"]; c.value = f"岗位：{JOB_TITLE}  |  部门：{JOB_DEPT}  |  生成日期：{JOB_DATE}"
c.fill = hfill("D6E4F0"); c.font = Font(size=11, color="1F4E79")
c.alignment = Alignment(horizontal="center", vertical="center"); ws1.row_dimensions[2].height = 24
meta = [
    ("岗位名称", JOB_TITLE),
    ("所在部门", JOB_DEPT),
    ("生成日期", JOB_DATE),
    ("能力维度数", f"{len(ABILITIES)} 项核心能力 + {len(TRAITS)} 项关键特质"),
    ("面试题总数", f"{len(QUESTIONS)} 道"),
    ("使用说明", "本表包含4个工作表：封面、能力模型、面试题库、打分表。\n面试时使用【打分表】为候选人评分，每位候选人建议单独复制一份使用。"),
]
for i, (lbl, val) in enumerate(meta, start=4):
    c = ws1.cell(row=i, column=1, value=lbl)
    c.fill = hfill("1F4E79"); c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = wa("center","center"); c.border = thin_border()
    cell(ws1, i, 2, val, bg="F2F2F2" if i % 2 == 0 else "FFFFFF")
    ws1.row_dimensions[i].height = 44 if i == 9 else 24

# Sheet2: 能力模型
ws2 = wb.create_sheet("能力模型")
for col, w in zip("ABCD", [12, 8, 18, 58]):
    ws2.column_dimensions[col].width = w
hdr(ws2, 1, range(1,5), ["类型","编号","名称","说明"])
ws2.row_dimensions[1].height = 24
row = 2
for code, name, desc in ABILITIES:
    cell(ws2, row, 1, "核心能力", bg="D6E4F0", bold=True, h="center")
    cell(ws2, row, 2, code,       bg="D6E4F0", bold=False, h="center")
    cell(ws2, row, 3, name,       bg="D6E4F0", bold=True)
    cell(ws2, row, 4, desc,       bg="D6E4F0")
    ws2.row_dimensions[row].height = 42; row += 1
for code, name, desc in TRAITS:
    cell(ws2, row, 1, "关键特质", bg="E2EFDA", bold=True, h="center")
    cell(ws2, row, 2, code,       bg="E2EFDA", bold=False, h="center")
    cell(ws2, row, 3, name,       bg="E2EFDA", bold=True)
    cell(ws2, row, 4, desc,       bg="E2EFDA")
    ws2.row_dimensions[row].height = 42; row += 1
ws2.freeze_panes = "A2"

# Sheet3: 面试题库
ws3 = wb.create_sheet("面试题库")
for i, w in enumerate([6,14,8,44,28,28,28,28,34,22,22,22,22,22], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
hdr(ws3, 1, range(1,15), [
    "编号","考察维度","题型","面试题目",
    "S-情境追问","T-任务追问","A-行动追问","R-结果追问","深挖追问",
    "1分（不达标）","2分（基本达标）","3分（达标）","4分（良好）","5分（优秀）"
])
ws3.row_dimensions[1].height = 24
for i, q in enumerate(QUESTIONS, start=2):
    bg = "D6E4F0" if q["dt"] == "核心能力" else "E2EFDA"
    for col, val in enumerate([
        q["id"],q["dim"],q["tp"],q["q"],
        q["sS"],q["sT"],q["sA"],q["sR"],q["sX"],
        q["s1"],q["s2"],q["s3"],q["s4"],q["s5"]
    ], 1):
        cell(ws3, i, col, val, bg=(bg if col <= 3 else None))
    ws3.row_dimensions[i].height = 80
ws3.freeze_panes = "D2"

# Sheet4: 打分表
ws4 = wb.create_sheet("打分表")
for col, w in zip("ABCDEF", [7, 44, 14, 10, 38, 16]):
    ws4.column_dimensions[col].width = w
for col, lbl in enumerate(["候选人姓名：","","面试日期：","","面试官：",""], 1):
    c = ws4.cell(row=1, column=col, value=lbl)
    c.fill = hfill("F2F2F2"); c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center"); c.border = thin_border()
ws4.row_dimensions[1].height = 26
ws4.row_dimensions[2].height = 6
hdr(ws4, 3, range(1,7), ["编号","面试题目（简版）","考察维度","评分（1-5）","关键行为记录","备注"])
ws4.row_dimensions[3].height = 24
for i, q in enumerate(QUESTIONS, start=4):
    short_q = (q["q"][:38] + "…") if len(q["q"]) > 38 else q["q"]
    bg = "D6E4F0" if q["dt"] == "核心能力" else "E2EFDA"
    cell(ws4, i, 1, q["id"], bg=bg, h="center")
    cell(ws4, i, 2, short_q)
    cell(ws4, i, 3, q["dim"], bg=bg, h="center")
    c = ws4.cell(row=i, column=4, value="")
    c.fill = hfill("FCE4D6")
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
    cell(ws4, i, 5, ""); cell(ws4, i, 6, "")
    ws4.row_dimensions[i].height = 44
total_row = len(QUESTIONS) + 4
for col in range(1, 7):
    c = ws4.cell(row=total_row, column=col)
    c.fill = hfill("FFF2CC"); c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
ws4.cell(total_row, 2).value = "总  分"
ws4.cell(total_row, 4).value = f"=SUM(D4:D{total_row-1})"
ws4.row_dimensions[total_row].height = 26
eval_row = total_row + 1
ws4.merge_cells(f"B{eval_row}:F{eval_row}")
for col in range(1, 7):
    c = ws4.cell(row=eval_row, column=col)
    c.fill = hfill("F2F2F2"); c.font = Font(bold=True, size=10)
    c.alignment = wa("left","center"); c.border = thin_border()
ws4.cell(eval_row, 2).value = "综合评价与录用建议："
ws4.row_dimensions[eval_row].height = 54
ws4.freeze_panes = "A4"

# ── 保存 ──
out = os.path.join(os.path.dirname(__file__), "高级产品经理-面试打分表-样例.xlsx")
wb.save(out)
print(f"✅ 已生成：{out}")
